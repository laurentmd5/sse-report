from datetime import datetime
from sqlalchemy import func
from app import db
from app.models.intervention import Intervention
from app.models.team import Team
from app.models.user import User
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

def get_monthly_stats(year, month):
    """
    Calcule les statistiques globales pour un mois donné.
    """
    # 1. Total des interventions validées ce mois-ci
    total_interventions = Intervention.query.filter(
        db.extract('year', Intervention.intervention_date) == year,
        db.extract('month', Intervention.intervention_date) == month,
    ).count()
    
    # 2. Statistiques par Tâche
    task_stats_query = db.session.query(
        Intervention.task_type, 
        func.count(Intervention.id)
    ).filter(
        db.extract('year', Intervention.intervention_date) == year,
        db.extract('month', Intervention.intervention_date) == month
    ).group_by(Intervention.task_type).all()
    
    task_stats = {task: count for task, count in task_stats_query}
    
    # 3. Statistiques par Équipe
    team_stats_query = db.session.query(
        Team.team_name,
        func.count(Intervention.id)
    ).join(User, User.team_id == Team.id)\
     .join(Intervention, Intervention.team_leader_id == User.id)\
     .filter(
        db.extract('year', Intervention.intervention_date) == year,
        db.extract('month', Intervention.intervention_date) == month
    ).group_by(Team.team_name).all()
    
    team_stats = {team: count for team, count in team_stats_query}
    
    # 4. Tableau croisé (Équipe vs Tâche) pour le rapport
    # On ne récupère que les équipes qui ont des interventions pour ce mois-ci
    active_team_ids = db.session.query(User.team_id)\
        .join(Intervention, Intervention.team_leader_id == User.id)\
        .filter(
            db.extract('year', Intervention.intervention_date) == year,
            db.extract('month', Intervention.intervention_date) == month
        ).distinct().all()
    
    active_team_ids = [t[0] for t in active_team_ids if t[0] is not None]
    teams = Team.query.filter(Team.id.in_(active_team_ids)).all() if active_team_ids else []
    
    cross_table = []
    
    # Liste dynamique des tâches qui ont été effectuées ce mois-ci
    distinct_tasks = db.session.query(Intervention.task_type).filter(
        db.extract('year', Intervention.intervention_date) == year,
        db.extract('month', Intervention.intervention_date) == month
    ).distinct().all()
    
    report_tasks = [t[0] for t in distinct_tasks if t[0]]
    
    for team in teams:
        row = {'team_name': team.team_name, 'tasks': {t: 0 for t in report_tasks}, 'total': 0}
        
        # Récupérer les interventions de cette équipe ce mois-ci
        interventions = db.session.query(Intervention.task_type, func.count(Intervention.id))\
            .join(User, Intervention.team_leader_id == User.id)\
            .filter(
                User.team_id == team.id,
                db.extract('year', Intervention.intervention_date) == year,
                db.extract('month', Intervention.intervention_date) == month
            ).group_by(Intervention.task_type).all()
            
        for task_type, count in interventions:
            if task_type in row['tasks']:
                row['tasks'][task_type] += count
                row['total'] += count
                
        cross_table.append(row)
        
    return {
        'total': total_interventions,
        'task_stats': task_stats,
        'team_stats': team_stats,
        'cross_table': cross_table,
        'report_tasks': report_tasks
    }

def generate_excel_report(year, month, month_str):
    """
    Génère un fichier Excel en mémoire avec le format de suivi brut.
    """
    # Récupérer toutes les interventions du mois
    interventions = db.session.query(
        Intervention, User, Team
    ).join(User, Intervention.team_leader_id == User.id)\
     .outerjoin(Team, User.team_id == Team.id)\
     .filter(
        db.extract('year', Intervention.intervention_date) == year,
        db.extract('month', Intervention.intervention_date) == month
     ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Suivi {month_str}"
    
    # Styles
    header_font = Font(bold=True) # Texte noir et en gras
    # Orange similaire à la capture d'écran
    header_fill = PatternFill(start_color="FFF28224", end_color="FFF28224", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # En-têtes du fichier de suivi
    headers = ['Date', 'ND', 'N° Demande', 'Nom du Client', 'Equipe', 'Tâches effectuer']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Remplissage des données
    for row_num, (inv, usr, team) in enumerate(interventions, 2):
        ws.cell(row=row_num, column=1, value=inv.intervention_date.strftime('%d/%m/%Y')).alignment = center_align
        ws.cell(row=row_num, column=2, value=inv.nd).alignment = center_align
        ws.cell(row=row_num, column=3, value=inv.demande_no).alignment = center_align
        ws.cell(row=row_num, column=4, value=inv.client_name).alignment = center_align
        
        team_name = team.team_name if team else usr.full_name
        ws.cell(row=row_num, column=5, value=team_name).alignment = center_align
        ws.cell(row=row_num, column=6, value=inv.task_type).alignment = center_align
    
    # Ajuster la largeur des colonnes
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width
        
    # Sauvegarder dans un buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer
